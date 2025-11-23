from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from io import BytesIO
from pathlib import Path
import os

# ==========================================
# SECCIÓN DE PERSONALIZACIÓN DE COLORES
# ==========================================
# Aquí puedes cambiar los códigos Hexadecimales de los colores del reporte.
# Puedes usar herramientas online como "Color Picker" para obtener el código HEX.
# NOTA: No incluyas el símbolo '#' al principio.

COLOR_HEADER_ORANGE = "ED7D31"   # Color principal (Naranja) - Usado en Títulos y Encabezados
COLOR_SUBHEADER_GREEN = "006100" # Color secundario (Verde Oscuro) - Usado en Subtítulos

# --- COLORES PARA LA FILA DE TOTALES ---
COLOR_TOTAL_BG = "F2F2F2"      # Fondo Gris muy claro
COLOR_TOTAL_TEXT = "006100"    # Texto Verde oscuro (Combina con el secundario)
COLOR_WHITE = "FFFFFF"         # Blanco (Texto sobre fondos oscuros)

def generar_excel_ventas_producto(items: List[Dict[str, Any]], filtros_info: str, periodo_texto: str) -> BytesIO:
    """
    Genera el archivo Excel de ventas.
    
    Args:
        items: Lista de diccionarios con los datos de ventas.
        filtros_info: Texto que describe los filtros aplicados (ej. "Sucursal Centro").
        periodo_texto: Texto del periodo (ej. "Enero 2023").
        
    Returns:
        BytesIO: El archivo Excel en memoria listo para descargar.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Ventas"

    # --- DEFINICIÓN DE ESTILOS (FUENTES Y BORDES) ---
    font_white_bold = Font(color=COLOR_WHITE, bold=True, size=12, name='Arial')
    font_header_table = Font(color=COLOR_WHITE, bold=True, size=11, name='Arial')
    font_normal = Font(name='Arial', size=10)
    font_bold = Font(bold=True, name='Arial')

    # Estilos para la fila de Totales
    font_total_label = Font(color=COLOR_TOTAL_TEXT, bold=True, size=12, name='Arial')
    font_total_value = Font(color=COLOR_TOTAL_TEXT, bold=True, name='Arial')
    
    # Rellenos (Fondo de celdas)
    fill_orange = PatternFill(start_color=COLOR_HEADER_ORANGE, end_color=COLOR_HEADER_ORANGE, fill_type="solid")
    fill_green = PatternFill(start_color=COLOR_SUBHEADER_GREEN, end_color=COLOR_SUBHEADER_GREEN, fill_type="solid")
    fill_total = PatternFill(start_color=COLOR_TOTAL_BG, end_color=COLOR_TOTAL_BG, fill_type="solid")

    border_table = Border(
        left=Side(style='thin', color='000000'), 
        right=Side(style='thin', color='000000'), 
        top=Side(style='thin', color='000000'), 
        bottom=Side(style='thin', color='000000')
    )

    # ==========================================
    # SECCIÓN DE ENCABEZADOS (FILAS 1-4)
    # ==========================================
    
    # Fila 1: Nombre de la Empresa
    ws.merge_cells('A1:G1')
    cell_a1 = ws['A1']
    cell_a1.value = "COMERCIALIZADORA LA TRASQUILA SA DE CV" # <--- CAMBIAR NOMBRE EMPRESA AQUÍ
    cell_a1.font = font_white_bold
    cell_a1.fill = fill_orange
    cell_a1.alignment = Alignment(horizontal='left', vertical='center')

    # Fila 2: RFC
    ws.merge_cells('A2:G2')
    cell_a2 = ws['A2']
    cell_a2.value = "RFC: CTR2506114T9" # <--- CAMBIAR RFC AQUÍ
    cell_a2.font = font_white_bold
    cell_a2.fill = fill_orange
    cell_a2.alignment = Alignment(horizontal='left', vertical='center')

    # Fila 3: Sucursal (Dinámico)
    ws.merge_cells('A3:G3')
    cell_a3 = ws['A3']
    cell_a3.value = f"SUCURSAL: {filtros_info}"
    cell_a3.font = font_white_bold
    cell_a3.fill = fill_green
    cell_a3.alignment = Alignment(horizontal='left', vertical='center')

    # Fila 4: Título del Reporte
    ws.merge_cells('A4:G4')
    cell_a4 = ws['A4']
    cell_a4.value = f"REPORTE DE VENTA POR PRODUCTO {periodo_texto}" 
    cell_a4.font = font_white_bold
    cell_a4.fill = fill_green
    cell_a4.alignment = Alignment(horizontal='left', vertical='center')

    # ==========================================
    # SECCIÓN DE LOGO
    # ==========================================
    # El logo se busca en la carpeta raíz del backend (junto a main.py o un nivel arriba).
    # Para cambiar el logo, simplemente reemplaza el archivo "logo.png" en esa carpeta.
    ws.merge_cells('H1:I4')
    
    base_path = Path(__file__).resolve().parent.parent.parent 
    ruta_logo = base_path / "logo.png" # <--- Nombre del archivo de imagen
    
    if ruta_logo.exists():
        try:
            img = Image(str(ruta_logo))
            # Ajusta el tamaño del logo aquí si es necesario
            img.height = 65
            img.width = 180 
            ws.add_image(img, 'H1')
        except Exception:
            pass

    ws.append([]) # Fila 5 vacía para separación

    # ==========================================
    # SECCIÓN DE COLUMNAS (TABLA)
    # ==========================================
    # Para agregar o quitar columnas:
    # 1. Modifica la lista 'headers' abajo.
    # 2. Modifica el bucle 'for item in items' más abajo para asignar el valor correcto.
    
    headers = ["Código", "Producto", "Cantidad", "Unidad", "Precio", "Importe", "Descuento", "Impuesto", "Total"]
    ws.append(headers)
    
    # Estilo de la fila de encabezados (Fila 6)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num)
        cell.font = font_header_table
        cell.fill = fill_orange
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(right=Side(style='thin', color='FFFFFF'))

    # --- LLENADO DE DATOS ---
    row_idx = 7
    
    # Acumuladores para los totales
    total_cantidad = 0.0
    total_importe = 0.0
    total_descuento = 0.0
    total_impuesto = 0.0
    total_general = 0.0

    for item in items:
        # Columna 1: Código
        ws.cell(row=row_idx, column=1, value=item.get("CCODIGOPRODUCTO")).font = font_normal
        
        # Columna 2: Nombre Producto
        ws.cell(row=row_idx, column=2, value=item.get("CNOMBREPRODUCTO")).font = font_normal
        
        # Columna 3: Cantidad
        cant = float(item.get("cantidad", 0))
        ws.cell(row=row_idx, column=3, value=cant).number_format = '#,##0.00'
        ws.cell(row=row_idx, column=3).font = font_normal

        # Columna 4: Unidad
        ws.cell(row=row_idx, column=4, value=item.get("CNOMBREUNIDAD")).font = font_normal
        
        # Columna 5: Precio Unitario
        precio = float(item.get("precio", 0))
        ws.cell(row=row_idx, column=5, value=precio).number_format = '$#,##0.00'

        # Columna 6: Importe (Subtotal)
        importe = float(item.get("Importe", 0))
        ws.cell(row=row_idx, column=6, value=importe).number_format = '$#,##0.00'

        # Columna 7: Descuento
        desc = float(item.get("descuento", 0))
        ws.cell(row=row_idx, column=7, value=desc).number_format = '$#,##0.00'

        # Columna 8: Impuesto
        imp = float(item.get("impuesto", 0))
        ws.cell(row=row_idx, column=8, value=imp).number_format = '$#,##0.00'

        # Columna 9: Total
        tot = float(item.get("Total", 0))
        c_tot = ws.cell(row=row_idx, column=9, value=tot)
        c_tot.number_format = '$#,##0.00'
        c_tot.font = font_bold

        # Aplicar bordes a todas las celdas de la fila
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).border = border_table

        # Sumar a los totales
        total_cantidad += cant
        total_importe += importe
        total_descuento += desc
        total_impuesto += imp
        total_general += tot

        row_idx += 1

    # ==========================================
    # SECCIÓN DE TOTALES FINALES
    # ==========================================
    
    # Etiqueta "Total General" (Fusionada A y B)
    ws.merge_cells(f'A{row_idx}:B{row_idx}')
    cell_total = ws.cell(row=row_idx, column=1, value="Total General")
    cell_total.font = font_total_label
    cell_total.fill = fill_total       
    cell_total.alignment = Alignment(horizontal='left')
    
    # Asegurar estilo en la celda fusionada B
    ws.cell(row=row_idx, column=2).fill = fill_total
    ws.cell(row=row_idx, column=2).border = border_table

    # Total Cantidad
    c_cant = ws.cell(row=row_idx, column=3, value=total_cantidad)
    c_cant.font = font_total_value
    c_cant.fill = fill_total       
    c_cant.number_format = '#,##0.00'
    c_cant.border = border_table

    # Celdas vacías (Unidad, Precio) - Solo fondo gris
    ws.cell(row=row_idx, column=4).fill = fill_total
    ws.cell(row=row_idx, column=4).border = border_table
    ws.cell(row=row_idx, column=5).fill = fill_total
    ws.cell(row=row_idx, column=5).border = border_table

    # Total Importe
    c_imp = ws.cell(row=row_idx, column=6, value=total_importe)
    c_imp.font = font_total_value
    c_imp.fill = fill_total
    c_imp.number_format = '$#,##0.00'
    c_imp.border = border_table

    # Total Descuento
    c_desc = ws.cell(row=row_idx, column=7, value=total_descuento)
    c_desc.font = font_total_value
    c_desc.fill = fill_total
    c_desc.number_format = '$#,##0.00'
    c_desc.border = border_table

    # Total Impuesto
    c_impu = ws.cell(row=row_idx, column=8, value=total_impuesto)
    c_impu.font = font_total_value
    c_impu.fill = fill_total
    c_impu.number_format = '$#,##0.00'
    c_impu.border = border_table

    # Total Final
    c_tot = ws.cell(row=row_idx, column=9, value=total_general)
    c_tot.font = font_total_value
    c_tot.fill = fill_total
    c_tot.number_format = '$#,##0.00'
    c_tot.border = border_table

    # --- AJUSTE DE ANCHO DE COLUMNAS ---
    # Define el ancho visual de cada columna (aprox. caracteres)
    column_widths = [15, 50, 12, 10, 12, 15, 12, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Activar filtros automáticos
    ws.auto_filter.ref = f"A6:I{row_idx-1}"
    
    # Congelar paneles (para que el encabezado siempre se vea al bajar)
    ws.freeze_panes = "A7"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output